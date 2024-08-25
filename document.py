import os
import pikepdf
import docx
import openpyxl
import pptx

# Dictionary to map file extensions to their corresponding types
FILE_EXTENSIONS = {
    '.docx': 'word',
    '.pdf': 'pdf',
    '.xlsx': 'excel',
    '.pptx': 'powerpoint'
}

class DocumentMetadata:
    def __init__(self, file_path: str):
        # Initialize the DocumentMetadata object with a file path
        self.file_path = file_path

    def extract_metadata(self) -> dict:
        # Extract metadata based on file extension
        file_extension = os.path.splitext(self.file_path)[1]
        if file_extension in FILE_EXTENSIONS:
            metadata_method = getattr(self, f"{FILE_EXTENSIONS[file_extension]}_metadata")
            try:
                return metadata_method()
            except Exception as e:
                return {"Error": f"Failed to extract metadata: {str(e)}"}
        else:
            return {"Error": f"Unsupported file type: {file_extension}"}


    def word_metadata(self) -> dict:
        try:
            doc = docx.Document(self.file_path)    
        except docx.exceptions.PackageNotFoundError:
            return {"Error": "Failed to open Word document"}
        except Exception as e:
            return {"Error": f"Failed to extract Word metadata: {str(e)}"}

        metadata = {}
        metadata['Author'] = doc.core_properties.author
        metadata['Title'] = doc.core_properties.title
        metadata['Subject'] = doc.core_properties.subject
        metadata['Date Created'] = doc.core_properties.created
        metadata['Date Modified'] = doc.core_properties.modified
        metadata['Page Count'] = len(doc.sections)
        metadata['Revision'] = doc.core_properties.revision
        metadata['Last Printed'] = doc.core_properties.last_printed
        metadata['Category'] = doc.core_properties.category
        metadata['Keywords'] = doc.core_properties.keywords
        metadata['Comments'] = doc.core_properties.comments
        metadata['Version'] = doc.core_properties.version
        metadata['Status'] = doc.core_properties.content_status
        

        word_count = 0
        for para in doc.paragraphs:
            word_count += len(para.text.split())
        metadata['Word Count'] = word_count
        return metadata

    def pdf_metadata(self) -> dict:
        try:
            with pikepdf.Pdf.open(self.file_path) as pdf:
                metadata = {}
                metadata['Author'] = pdf.docinfo.get('/Author', '')
                metadata['Title'] = pdf.docinfo.get('/Title', '')
                metadata['Date Created'] = pdf.docinfo.get('/CreationDate', '')
                metadata['Date Modified'] = pdf.docinfo.get('/ModDate', '')
                metadata['Subject'] = pdf.docinfo.get('/Subject', '')
                metadata['Keywords'] = pdf.docinfo.get('/Keywords', '')
                metadata['PDDocID'] = pdf.docinfo.get('/ID', '')
                metadata['PDFVersion'] = pdf.pdf_version
                metadata['Page Count'] = len(pdf.pages)
                return metadata  
        except pikepdf.PdfError:
            return {"Error": "Failed to open PDF"}
        except Exception as e:
            return {"Error": f"Failed to extract PDF metadata: {str(e)}"}

    def excel_metadata(self) -> dict: 
        try:
            excel = openpyxl.load_workbook(self.file_path)   
        except openpyxl.exceptions.InvalidFileException:
            return {"Error": "Failed to open Excel file"}
        except Exception as e:
            return {"Error": f"Failed to extract Excel metadata: {str(e)}"}

        metadata = {}
        metadata['Author'] = excel.properties.creator
        metadata['Title'] = excel.properties.title
        metadata['Subject'] = excel.properties.subject
        metadata['Date Created'] = excel.properties.created
        metadata['Date Modified'] = excel.properties.modified
        metadata['Sheet Count'] = len(excel.worksheets)
        metadata['Row Count'] = excel.active.max_row
        metadata['Column Count'] = excel.active.max_column
        
        return metadata

    def powerpoint_metadata(self) -> dict:
        try:
            ppt = pptx.Presentation(self.file_path)
        except pptx.exceptions.PackageNotFoundError:
            return {"Error": "Failed to open PowerPoint presentation"}
        except Exception as e:
            return {"Error": f"Failed to extract PowerPoint metadata: {str(e)}"}
            
        metadata = {}
        metadata['Author'] = ppt.core_properties.author
        metadata['Title'] = ppt.core_properties.title
        metadata['Subject'] = ppt.core_properties.subject
        metadata['Date Created'] = ppt.core_properties.created
        metadata['Date Modified'] = ppt.core_properties.modified
        metadata['Slide Count'] = len(ppt.slides)
        metadata['Revision'] = ppt.core_properties.revision
        metadata['Last Printed'] = ppt.core_properties.last_printed
        metadata['Category'] = ppt.core_properties.category
        metadata['Keywords'] = ppt.core_properties.keywords
        metadata['Comments'] = ppt.core_properties.comments
        metadata['Version'] = ppt.core_properties.version
        metadata['Status'] = ppt.core_properties.content_status
        return metadata